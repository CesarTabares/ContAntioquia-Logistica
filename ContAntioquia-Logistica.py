# -*- coding: utf-8 -*-
"""
Created on Sat Feb  1 07:28:37 2020

@author: Cesar
"""

from datetime import datetime
from copy import deepcopy

import wx
import openpyxl
import smtplib
from email.message import EmailMessage


col_requerimiento_auto=1
col_area_req=2
col_area=8
col_fecha_auto=3
col_cotizacion=4
col_nombrecliente=5
col_tipo_req=6
col_tipotransporte=7
col_tipocontenedor=9
col_requieredescargue=10
col_origen=11
col_destino=12
col_km=13
col_precio=14
col_recargopeaje=15
col_precio_recargo=16
col_nombreresponsable=17
col_telefono_resp=18
col_cargo=19
col_nombresiso=20
col_telefono_siso=21
col_debeinfo=22
col_horasantes=23
col_fechaentrega=24
col_direccion=25
col_referenciacont=26
col_nombreconduc=27
col_cedula=28
col_telefonoconduc=29
col_placa=30
col_adiciones=31
col_preguntahoras=32
col_preguntadoc=33


principal_color=wx.Colour(51, 102, 51)
secondary_color='white'
yellow_color=(255, 203, 27)

path_config='Config.xlsx'
wb_config=openpyxl.load_workbook('Config.xlsx')
sheet_config=wb_config['Config']
path_db=sheet_config.cell(row=2,column=2).value
path_remision=sheet_config.cell(row=3,column=2).value
path_remision_nuevas=sheet_config.cell(row=4,column=2).value



class MyFrame(wx.Frame):
    
    
    def OnKeyDown(self, event):
        """quit if user press q or Esc"""
        if event.GetKeyCode() == 27 or event.GetKeyCode() == ord('Q'): #27 is Esc
            self.Close(force=True)
            
        else:
            event.Skip()
    
    def __init__(self):
        
        wx.Frame.__init__(self, None, wx.ID_ANY, "Centro Logistico", size=(575,290),style=wx.DEFAULT_FRAME_STYLE & ~(wx.RESIZE_BORDER | wx.MAXIMIZE_BOX))  
        self.Bind(wx.EVT_KEY_UP, self.OnKeyDown)
        self.SetBackgroundColour(secondary_color)
        self.panel = MainPanel(self)
        panel_font= wx.Font(10, wx.DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Normal")
        self.panel.SetFont(panel_font)
        
        
        self.Center()
        ico = wx.Icon('Cont.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        self.fgs= wx.GridBagSizer(0,0)
        
        title_font= wx.Font(15, wx.DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Normal")
        

        self.lbltitle =wx.StaticText(self.panel, label='Bienvenido al Centro Logistico de Contenedores de Antioquia')
        self.lbltitle.SetFont(title_font)
        
        self.lbltitle.SetBackgroundColour(secondary_color)
        self.lbltitle.SetForegroundColour(principal_color)
        self.fgs.Add(self.lbltitle,pos=(6,1),span=(1,5), flag=wx.LEFT | wx.ALIGN_CENTER, border=10)
        
        self.lbltitle2 =wx.StaticText(self.panel, label=' Que Desea Hacer ?')
        self.lbltitle2.SetFont(title_font)
        self.lbltitle2.SetBackgroundColour(secondary_color)
        self.lbltitle2.SetForegroundColour(principal_color)
        self.fgs.Add(self.lbltitle2,pos=(7,1),span=(1,5), flag=wx.LEFT | wx.ALIGN_CENTER, border=10)
            
        btn_nuevo_req = wx.Button(self.panel, id=wx.ID_ANY,  size=(100,40), label="Nuevo\nRequerimiento")
        self.fgs.Add(btn_nuevo_req, pos=(9,2),span=(1,1),flag= wx.RIGHT| wx.ALIGN_CENTER| wx.EXPAND, border=20)
        btn_nuevo_req.Bind(wx.EVT_BUTTON, self.open_nuevo_req11)
        
        btn_logistico = wx.Button(self.panel, id=wx.ID_ANY,size=(100,40), label="Logistica")
        self.fgs.Add(btn_logistico, pos=(9,3),span=(1,1), flag= wx.RIGHT | wx.ALIGN_CENTER| wx.EXPAND, border=20)
        btn_logistico.Bind(wx.EVT_BUTTON, self.open_logistica21)
        
        btn_imprimir = wx.Button(self.panel, id=wx.ID_ANY, size=(100,40),label="Imprimir\nRemision")
        self.fgs.Add(btn_imprimir, pos=(9,4),span=(1,1), flag= wx.RIGHT | wx.ALIGN_CENTER| wx.EXPAND, border=0)
        btn_imprimir.Bind(wx.EVT_BUTTON, self.open_imprimir11)
        
        #btn_logistico = wx.Button(self.panel, id=wx.ID_ANY, label="Configuracion",size=(-1,-1))
        #self.fgs.Add(btn_logistico, pos=(17,6),span=(1,1), flag= wx.ALL, border=0)
        #btn_logistico.Bind(wx.EVT_BUTTON, self.configuracion)
        
        
        
        mainSizer= wx.BoxSizer(wx.VERTICAL)
        mainSizer.Add(self.fgs,0, flag=wx.ALIGN_LEFT)
        self.panel.SetSizerAndFit(mainSizer)
            
    #-------------Button Functions-----------------#
    def open_nuevo_req11(self, event):
        ww_nuevo_requerimiento11(parent=self.panel).Show()
       

    def open_logistica21(self, event):
        ww_logistica21(parent=self.panel).Show()
    
    def open_imprimir11(self,evento):
        ww_remision11(parent=self.panel).Show()
        
    def configuracion(self, event):
        ww_configuracion(parent=self.panel).Show()

    #-------------Button Functions-----------------#

class MainPanel(wx.Panel):

    def __init__(self,parent):
        # create the panel
        wx.Panel.__init__(self, parent=parent)
        try:

            image_file = 'LOGOpng-01-100.png'
            bmp1 = wx.Image(
                image_file, 
                wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            # image's upper left corner anchors at panel 
            # coordinates (0, 0)
            self.bitmap1 = wx.StaticBitmap(
                self, -1, bmp1, (190, 5))
            # show some image details
            #str1 = "%s  %dx%d" % (image_file, bmp1.GetWidth(),
                                  #bmp1.GetHeight()) 
            #parent.SetTitle(str1)
        except IOError:
            print ("Image file %s not found")
            raise SystemExit        
#https://stackoverflow.com/questions/15861288/set-picture-as-background
  
class ww_nuevo_requerimiento11(wx.Frame):   
    
    def __init__(self,parent):
        ######----------------------------------------BACK END----------------------------------------#############        
    
        wb_listas=openpyxl.load_workbook(path_config)
        
        
        
        req1_sheet=wb_listas['Requerimientos-1']
        
        areas=[]
        
        for cell in req1_sheet['A']:
            if cell.value != None:
                areas.append(cell.value)
        areas.pop(0)
        
        ######----------------------------------------BACK END----------------------------------------#############       
        
        ######----------------------------------------FRONT END----------------------------------------#############
        
        wx.Frame.__init__(self, None, wx.ID_ANY, "Contenedores de Antioquia - Centro Logistico", size=(250, 250),style=wx.DEFAULT_FRAME_STYLE & ~(wx.RESIZE_BORDER | wx.MAXIMIZE_BOX))  
        self.SetBackgroundColour(secondary_color)
        self.Center()
        try:
            
            #image_file = 'CINCO CONSULTORES.jpg'
            #bmp1 = wx.Image(
                #image_file, 
                #wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            #self.panel = wx.StaticBitmap(
                #self, -1, bmp1, (0, 0)
            self.panel=wx.Panel(self)
            panel_font= wx.Font(10, wx.DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Normal")
            self.panel.SetFont(panel_font)
            self.panel.SetBackgroundColour(secondary_color)

        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
        
        ico = wx.Icon('Cont.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        self.fgs= wx.GridBagSizer(0,0)
        
        title_font= wx.Font(11, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Normal")
       
        self.lbltitle =wx.StaticText(self.panel, label='Nuevo Requerimiento Por:')
        self.lbltitle.SetFont(title_font)
        self.lbltitle.SetBackgroundColour(secondary_color)
        self.lbltitle.SetForegroundColour(principal_color)
        self.fgs.Add(self.lbltitle,pos=(2,1),span=(1,3), flag=wx.ALL | wx.ALIGN_CENTER, border=5)

        self.combo_area = wx.ComboBox(self.panel,value=areas[0], choices=areas)
        self.fgs.Add(self.combo_area , pos=(4,1),span=(1,3), flag= wx.ALL |wx.ALIGN_CENTER, border=5)
        
        btn_aceptar = wx.Button(self.panel, id=wx.ID_ANY, label="Aceptar",size=(-1,-1))
        self.fgs.Add(btn_aceptar, pos=(6,1),span=(1,3), flag= wx.ALL | wx.ALIGN_CENTER, border=0)
        btn_aceptar.Bind(wx.EVT_BUTTON, self.open_nuevo_req12)

        mainSizer= wx.BoxSizer(wx.VERTICAL)
        mainSizer.Add(self.fgs,0, flag=wx.ALIGN_CENTER)
        self.panel.SetSizerAndFit(mainSizer)
        
    #-------------Button Functions-----------------#            
    def open_nuevo_req12(self, event):
        
        #try:
        
        self.Destroy()
        area_req=self.combo_area.GetValue()
        #pub.sendMessage("panel_listener",message=area_req,parent=self.panel)

        ww_nuevo_requerimiento12(parent=self.panel,message=area_req).Show()
        
        
        #except:
         #   error_msgbox=wx.MessageDialog(None,'Error al guardar el registro en la BD. \nVerifique el el archivo de excel este cerrado y en la ruta correcta.','ERROR',wx.ICON_ERROR)
          #  error_msgbox.ShowModal()
    #-------------Button Functions-----------------# 
        





class ww_nuevo_requerimiento12(wx.Frame):


        
    def __init__(self,parent,message):
        wb_req=openpyxl.load_workbook(path_db)
        #pub.subscribe(self.__init__, "panel_listener")
        self.area_selec=message
        
        wb_listas=openpyxl.load_workbook(path_config)
        req2_sheet=wb_listas['Requerimientos-12']
        hist_req_sheet=wb_req['Requerimientos']

        self.lista_encargado=[]
        self.lista_tipo_cont=[]
        self.lista_tipo_transp=[]
        self.lista_descargue=[]
        self.lista_debe_enviarinfo=[]
        self.lista_nro_req=[]
        self.lista_tiporeq=[]
        
        for cell in req2_sheet['A']:
            if cell.value != None:
                self.lista_encargado.append(cell.value)
        for cell in req2_sheet['B']:
            if cell.value != None:
                self.lista_tipo_cont.append(cell.value)
        for cell in req2_sheet['C']:
            if cell.value != None:
                self.lista_tipo_transp.append(cell.value)
        for cell in req2_sheet['E']:
            if cell.value != None:
                self.lista_descargue.append(cell.value)   
        for cell in req2_sheet['F']:
            if cell.value != None:
                self.lista_debe_enviarinfo.append(cell.value)
        for cell in req2_sheet['G']:
            if cell.value != None:
                self.lista_tiporeq.append(cell.value)
        
        
        for cell in hist_req_sheet['A']:
            if cell.value !=None:
                self.lista_nro_req.append(cell.value)
        
        try:
            self.nro_req= int(self.lista_nro_req[-1])+1
        except:
            self.nro_req=1

                
        self.lista_tipo_cont.pop(0)
        self.lista_tipo_transp.pop(0)
        self.lista_descargue.pop(0)
        self.lista_debe_enviarinfo.pop(0)
        self.lista_tiporeq.pop(0)
        
        self.fila_vacia = 2
        
        while (hist_req_sheet.cell(row = self.fila_vacia, column = 1).value != None) :
          self.fila_vacia += 1
 
        ######----------------------------------------BACK END----------------------------------------#############       
        
        ######----------------------------------------FRONT END----------------------------------------#############
        
        wx.Frame.__init__(self, None, wx.ID_ANY, "Centro Logistico", size=(930, 570),style=wx.DEFAULT_FRAME_STYLE & ~(wx.RESIZE_BORDER | wx.MAXIMIZE_BOX))  
        self.SetBackgroundColour(secondary_color)
        self.panel = NuevoReqPanel(self)
        panel_font= wx.Font(10, wx.DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Normal")
        self.panel.SetFont(panel_font)
        self.Center()
                
        ico = wx.Icon('Cont.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        self.fgs= wx.GridBagSizer(0,0)
        title_font= wx.Font(25, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Bold")
        title_font3= wx.Font(15, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Bold")
        bold_font= wx.Font(10, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Bold")
        
        self.lbltitle2 =wx.StaticText(self.panel, label='CENTRO LOGISTICO')
        self.lblrequerimiento =wx.StaticText(self.panel, label='Requerimiento N° ' + str(self.nro_req))
        self.requerimiento_auto =(self.nro_req)
        self.lblfecha =wx.StaticText(self.panel, label='Fecha')
        self.lblfecha_auto =wx.StaticText(self.panel, label=datetime.today().strftime('%d-%m-%Y')) #-%H:%M:%S
        self.lblarea_req =wx.StaticText(self.panel, label='Req. Por: ')
        self.lblarea_req_auto =wx.StaticText(self.panel, label=message)
        self.lblcotizacion =wx.StaticText(self.panel, label='Cotizacion N°')
        self.lbltipotransporte =wx.StaticText(self.panel, label='Tipo de Transporte')
        self.lbltipocontenedor =wx.StaticText(self.panel, label='Tipo de Contenedor')
        self.lblrequieredescargue =wx.StaticText(self.panel, label='Requiere Descargue')
        self.lblorigen =wx.StaticText(self.panel, label='Origen')
        self.lbldestino =wx.StaticText(self.panel, label='Destino')
        self.lblkm =wx.StaticText(self.panel, label='Km')
        self.lblprecio =wx.StaticText(self.panel, label='Precio')
        self.lblrecargopeaje =wx.StaticText(self.panel, label='Recargo Peaje')
        self.lblinfocliente =wx.StaticText(self.panel, label='INFORMACION CLIENTE')
        self.lblnombreresponsable =wx.StaticText(self.panel, label='Nombre Responsable')
        self.lbltelefono_resp =wx.StaticText(self.panel, label='Telefono')
        self.lblcargo =wx.StaticText(self.panel, label='Cargo')
        self.lblnombresiso =wx.StaticText(self.panel, label='Nombre SISO')
        self.lbltelefono_siso =wx.StaticText(self.panel, label='Telefono')
        self.lbldebeinfo =wx.StaticText(self.panel, label='Debe Enviarse\nInformacion')
        self.lblhorasantes =wx.StaticText(self.panel, label='N° Horas Antes')
        self.lbltiporeq =wx.StaticText(self.panel, label='Tipo Requerimiento')
        self.lblnombrecliente =wx.StaticText(self.panel, label='Nombre Cliente')
        
        self.lbltitle2.SetFont(title_font)
        self.lblinfocliente.SetFont(title_font3)
        
        self.lblrequerimiento.SetFont(bold_font)
        self.lblfecha.SetFont(bold_font)
        self.lblfecha_auto.SetFont(bold_font)
        self.lblcotizacion.SetFont(bold_font)
        self.lbltipotransporte.SetFont(bold_font)
        self.lbltipocontenedor.SetFont(bold_font)
        self.lblrequieredescargue.SetFont(bold_font)
        self.lblorigen.SetFont(bold_font)
        self.lbldestino.SetFont(bold_font)
        self.lblkm.SetFont(bold_font)
        self.lblprecio.SetFont(bold_font)
        self.lblrecargopeaje.SetFont(bold_font)
        self.lblnombreresponsable.SetFont(bold_font)
        self.lbltelefono_resp.SetFont(bold_font)
        self.lblcargo.SetFont(bold_font)
        self.lblnombresiso.SetFont(bold_font)
        self.lbltelefono_siso.SetFont(bold_font)
        self.lbldebeinfo.SetFont(bold_font)
        self.lblhorasantes.SetFont(bold_font)
        self.lbltiporeq.SetFont(bold_font)
        self.lblnombrecliente.SetFont(bold_font)
        
        self.lblrequerimiento.SetFont(title_font3)
        self.lblfecha.SetFont(title_font3)
        self.lblfecha_auto.SetFont(title_font3)
        self.lblarea_req.SetFont(bold_font)
        self.lblarea_req_auto.SetFont(bold_font)
        
        
        self.lbltitle2.SetBackgroundColour(secondary_color)
        self.lblrequerimiento.SetBackgroundColour(secondary_color)
        self.lblfecha.SetBackgroundColour(secondary_color)
        self.lblfecha_auto.SetBackgroundColour(secondary_color)
        self.lblarea_req.SetBackgroundColour(secondary_color)
        self.lblarea_req_auto.SetBackgroundColour(secondary_color)
        self.lblcotizacion.SetBackgroundColour(secondary_color)
        self.lbltipotransporte.SetBackgroundColour(secondary_color)
        self.lbltipocontenedor.SetBackgroundColour(secondary_color)
        self.lblrequieredescargue.SetBackgroundColour(secondary_color)
        self.lblorigen.SetBackgroundColour(secondary_color)
        self.lbldestino.SetBackgroundColour(secondary_color)
        self.lblkm.SetBackgroundColour(secondary_color)
        self.lblprecio.SetBackgroundColour(secondary_color)
        self.lblrecargopeaje.SetBackgroundColour(secondary_color)
        self.lblinfocliente.SetBackgroundColour(secondary_color)
        self.lblnombreresponsable.SetBackgroundColour(secondary_color)
        self.lbltelefono_resp.SetBackgroundColour(secondary_color)
        self.lblcargo.SetBackgroundColour(secondary_color)
        self.lblnombresiso.SetBackgroundColour(secondary_color)
        self.lbltelefono_siso.SetBackgroundColour(secondary_color)
        self.lbldebeinfo.SetBackgroundColour(secondary_color)
        self.lblhorasantes.SetBackgroundColour(secondary_color)
        
        
        self.lbltitle2.SetForegroundColour(principal_color)
        self.lblrequerimiento.SetForegroundColour(principal_color)
        self.lblfecha.SetForegroundColour(principal_color)
        self.lblfecha_auto.SetForegroundColour(principal_color)
        self.lblarea_req.SetForegroundColour(principal_color)
        self.lblarea_req_auto.SetForegroundColour(principal_color)
        self.lblcotizacion.SetForegroundColour(principal_color)
        self.lbltipotransporte.SetForegroundColour(principal_color)
        self.lbltipocontenedor.SetForegroundColour(principal_color)
        self.lblrequieredescargue.SetForegroundColour(principal_color)
        self.lblorigen.SetForegroundColour(principal_color)
        self.lbldestino.SetForegroundColour(principal_color)
        self.lblkm.SetForegroundColour(principal_color)
        self.lblprecio.SetForegroundColour(principal_color)
        self.lblrecargopeaje.SetForegroundColour(principal_color)
        self.lblinfocliente.SetForegroundColour(principal_color)
        self.lblnombreresponsable.SetForegroundColour(principal_color)
        self.lbltelefono_resp.SetForegroundColour(principal_color)
        self.lblcargo.SetForegroundColour(principal_color)
        self.lblnombresiso.SetForegroundColour(principal_color)
        self.lbltelefono_siso.SetForegroundColour(principal_color)
        self.lbldebeinfo.SetForegroundColour(principal_color)
        self.lblhorasantes.SetForegroundColour(principal_color)
        self.lbltiporeq.SetForegroundColour(principal_color)
        self.lblnombrecliente.SetForegroundColour(principal_color)
        
        self.txtcotizacion=wx.TextCtrl(self.panel)
        self.txtorigen=wx.TextCtrl(self.panel)
        self.txtdestino=wx.TextCtrl(self.panel)
        self.txtkm=wx.TextCtrl(self.panel)
        self.txtprecio=wx.TextCtrl(self.panel)
        self.txtnombreresponsable=wx.TextCtrl(self.panel)
        self.txttelefono_resp=wx.TextCtrl(self.panel)
        self.txtcargo=wx.TextCtrl(self.panel)
        self.txtnombresiso=wx.TextCtrl(self.panel)
        self.txttelefono_siso=wx.TextCtrl(self.panel)
        self.txthorasantes=wx.TextCtrl(self.panel)
        self.txtnombrecliente=wx.TextCtrl(self.panel)
        
        self.combotipotransporte=wx.ComboBox(self.panel,value=self.lista_tipo_transp[0], choices=self.lista_tipo_transp)
        self.combotipocontenedor=wx.ComboBox(self.panel,value=self.lista_tipo_cont[0], choices=self.lista_tipo_cont)
        self.comborequieredescargue=wx.ComboBox(self.panel,value=self.lista_descargue[0], choices=self.lista_descargue)
        self.combotiporeq=wx.ComboBox(self.panel,value=self.lista_tiporeq[0], choices=self.lista_tiporeq)
        
        self.check_si_peaje = wx.CheckBox(self.panel, label= "Si")
        self.check_no_peaje = wx.CheckBox(self.panel, label='No')
        self.check_si_info = wx.CheckBox(self.panel, label= "Si")
        self.check_no_info = wx.CheckBox(self.panel, label='No')
        
        self.check_si_peaje.SetForegroundColour(principal_color)
        self.check_no_peaje.SetForegroundColour(principal_color)
        self.check_si_info.SetForegroundColour(principal_color)
        self.check_no_info.SetForegroundColour(principal_color)
        
        self.check_no_peaje.SetValue(True)

        btn_guardar = wx.Button(self.panel, id=wx.ID_OK, label="Guardar",size=(-1,-1))
        btn_salir = wx.Button(self.panel, id=wx.ID_ANY, label="Salir",size=(-1,-1))
        btn_adicionar_transp = wx.Button(self.panel, id=wx.ID_OK, label="Adicionar",size=(-1,-1))
        
        order=(self.txtcotizacion,self.txtnombrecliente, self.combotiporeq,self.combotipotransporte,self.combotipocontenedor,self.comborequieredescargue,self.txtorigen,self.txtdestino,self.txtkm,self.txtprecio,
               self.check_si_peaje,self.check_no_peaje,self.txtnombreresponsable,self.txttelefono_resp,self.txtcargo,self.txtnombresiso,self.txttelefono_siso,self.check_si_info,
               self.check_no_info,self.txthorasantes,btn_adicionar_transp,btn_guardar,btn_salir)
        
        for i in range(len(order) - 1):
            order[i+1].MoveAfterInTabOrder(order[i])
        
        self.fgs.Add(self.check_si_peaje, pos=(8,6),span=(1,1), flag= wx.ALL  |wx.ALIGN_RIGHT, border=5)
        self.fgs.Add(self.check_no_peaje, pos=(8,7),span=(1,1), flag= wx.ALL |wx.ALIGN_LEFT, border=5)
        self.fgs.Add(self.check_si_info, pos=(15,2),span=(1,1), flag= wx.ALL  |wx.ALIGN_LEFT, border=5)
        self.fgs.Add(self.check_no_info, pos=(16,2),span=(1,1), flag= wx.LEFT |wx.ALIGN_LEFT, border=5)
        
        self.fgs.Add(btn_adicionar_transp, pos=(17,6),span=(1,1), flag= wx.ALL | wx.ALIGN_CENTER, border=5)
        self.fgs.Add(btn_guardar, pos=(17,7),span=(1,1), flag= wx.ALL | wx.ALIGN_CENTER, border=5)
        self.fgs.Add(btn_salir, pos=(17,8),span=(1,1), flag= wx.ALL | wx.ALIGN_CENTER, border=5)
               
        self.fgs.Add(self.combotipotransporte,pos=(6,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.combotipocontenedor,pos=(7,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.comborequieredescargue,pos=(8,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.combotiporeq, pos=(5,2),span=(1,1), flag= wx.ALL, border=5)

        self.fgs.Add(self.txtcotizacion, pos=(4,2),span=(1,1), flag= wx.ALL , border=5)
        self.fgs.Add(self.txtnombrecliente, pos=(4,5),span=(1,2), flag= wx.ALL| wx.EXPAND, border=5)
        self.fgs.Add(self.txtorigen, pos=(6,4),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtdestino, pos=(6,5),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtkm, pos=(6,6),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtprecio, pos=(8,5),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtnombreresponsable, pos=(12,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txttelefono_resp, pos=(12,5),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtcargo, pos=(12,8),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtnombresiso, pos=(14,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txttelefono_siso, pos=(14,5),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txthorasantes, pos=(17,2),span=(1,1), flag= wx.ALL, border=5)

        self.fgs.Add(self.lblnombrecliente , pos=(4,4),span=(1,1), flag= wx.ALL | wx.ALIGN_CENTER, border=5)
        self.fgs.Add(self.lbltitle2 , pos=(1,1),span=(1,8), flag= wx.ALL | wx.ALIGN_CENTER, border=5)
        self.fgs.Add(self.lblrequerimiento , pos=(2,1),span=(1,2), flag= wx.ALL|wx.ALIGN_BOTTOM, border=5)
        self.fgs.Add(self.lblfecha , pos=(2,7),span=(1,1), flag= wx.ALL | wx.ALIGN_RIGHT |wx.ALIGN_BOTTOM, border=0)
        self.fgs.Add(self.lblfecha_auto , pos=(2,8),span=(1,1), flag= wx.LEFT|wx.ALIGN_BOTTOM, border=5)
        self.fgs.Add(self.lblarea_req , pos=(3,7),span=(1,1), flag= wx.ALL|wx.ALIGN_TOP | wx.ALIGN_RIGHT, border=0)
        self.fgs.Add(self.lblarea_req_auto , pos=(3,8),span=(1,1), flag= wx.LEFT |wx.ALIGN_TOP, border=5)
        self.fgs.Add(self.lblcotizacion , pos=(4,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbltiporeq , pos=(5,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbltipotransporte , pos=(6,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbltipocontenedor , pos=(7,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblrequieredescargue, pos=(8,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblorigen , pos=(5,4),span=(1,1), flag= wx.ALL | wx.ALIGN_CENTER, border=0)
        self.fgs.Add(self.lbldestino , pos=(5,5),span=(1,1), flag= wx.ALL| wx.ALIGN_CENTER, border=0)
        self.fgs.Add(self.lblkm , pos=(5,6),span=(1,1), flag= wx.ALL| wx.ALIGN_CENTER, border=0)
        self.fgs.Add(self.lblprecio , pos=(7,5),span=(1,1), flag= wx.ALL |wx.ALIGN_BOTTOM | wx.ALIGN_CENTER_HORIZONTAL, border=0)
        self.fgs.Add(self.lblrecargopeaje , pos=(7,6),span=(1,2), flag= wx.LEFT |wx.ALIGN_BOTTOM | wx.ALIGN_CENTER_HORIZONTAL, border=11)
        self.fgs.Add(self.lblinfocliente , pos=(10,1),span=(1,8), flag= wx.ALL| wx.ALIGN_CENTER, border=5)
        self.fgs.Add(self.lblnombreresponsable , pos=(12,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbltelefono_resp , pos=(12,4),span=(1,1), flag= wx.ALL | wx.ALIGN_RIGHT, border=5)
        self.fgs.Add(self.lblcargo , pos=(12,7),span=(1,1), flag= wx.ALL| wx.ALIGN_RIGHT, border=5)
        self.fgs.Add(self.lblnombresiso , pos=(14,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbltelefono_siso , pos=(14,4),span=(1,1), flag= wx.ALL |wx.ALIGN_RIGHT, border=5)
        self.fgs.Add(self.lbldebeinfo , pos=(15,1),span=(2,1), flag= wx.ALL| wx.ALIGN_CENTER_VERTICAL, border=5)
        self.fgs.Add(self.lblhorasantes , pos=(17,1),span=(1,1), flag= wx.ALL, border=5)
        
        
        self.check_si_peaje.Bind(wx.EVT_CHECKBOX, self.onCheck_si_peaje)
        self.check_no_peaje.Bind(wx.EVT_CHECKBOX, self.onCheck_no_peaje)
        self.check_si_info.Bind(wx.EVT_CHECKBOX, self.onCheck_si_info)
        self.check_no_info.Bind(wx.EVT_CHECKBOX, self.onCheck_no_info)
        
        btn_guardar.Bind(wx.EVT_BUTTON, self.guardar_req)
        btn_salir.Bind(wx.EVT_BUTTON, self.salir)
        btn_adicionar_transp.Bind(wx.EVT_BUTTON, self.adicionar_transp)
        
        mainSizer= wx.BoxSizer(wx.VERTICAL)
        mainSizer.Add(self.fgs,0, flag=wx.ALIGN_LEFT)
        self.panel.SetSizerAndFit(mainSizer) 
        
        self.txthorasantes.Hide()
        self.lblhorasantes.Hide()
        
    def onCheck_si_peaje(self,event):
        if self.check_no_peaje.IsChecked():
            self.check_no_peaje.SetValue(False)
            
    def onCheck_no_peaje(self,event):
        if self.check_si_peaje.IsChecked():
            self.check_si_peaje.SetValue(False)
            
    def onCheck_si_info(self,event):
        self.txthorasantes.Show()
        self.lblhorasantes.Show()
        
        if self.check_no_info.IsChecked():
            self.check_no_info.SetValue(False)
            
    def onCheck_no_info(self,event):
        self.txthorasantes.Hide()
        self.lblhorasantes.Hide()
        
        if self.check_si_info.IsChecked():
            self.check_si_info.SetValue(False)
    
    def precio_final(self,hist_req_sheet):
        wb_listas=openpyxl.load_workbook(path_config)
        config_sheet=wb_listas['Config']
        
        valor_recargo=config_sheet.cell(row=1,column=2).value
        if self.check_si_peaje.IsChecked():
            
            hist_req_sheet.cell(row=self.fila_vacia, column=col_precio_recargo).value=int(self.txtprecio.GetValue())+valor_recargo
        else:
            hist_req_sheet.cell(row=self.fila_vacia, column=col_precio_recargo).value=int(self.txtprecio.GetValue())

    def guardar_req(self,event):
        
        
        if self.check_si_info.IsChecked():
            diccionario_campos_oblig_texto={self.txtcotizacion:'Cotizacion N°', self.txtnombrecliente:'Nombre Cliente', self.txtorigen:'Origen', self.txtdestino:'Destino', self.txtnombreresponsable:'Nombre Responsable', self.txttelefono_resp:'Telefono Resposnable', self.txthorasantes:'N° Horas Antes'}
        else:
            diccionario_campos_oblig_texto={self.txtcotizacion:'Cotizacion N°', self.txtnombrecliente:'Nombre Cliente',self.txtorigen:'Origen', self.txtdestino:'Destino', self.txtnombreresponsable:'Nombre Responsable', self.txttelefono_resp:'Telefono Resposnable'}
        
        diccionario_campos_oblig_numero={self.txtprecio:'Precio'}
        diccionario_campos_oblig_combos={self.combotipotransporte:'Tipo de Transporte', self.combotiporeq:'Tipo Requerimiento', self.combotipocontenedor:'Tipo de Contenedor', self.comborequieredescargue:'Requiere Desacargue'} 
        
        if self.validar_campos_vacios_texto(diccionario_campos_oblig_texto)==False or self.validar_campos_vacios_numero(diccionario_campos_oblig_numero)==False or self.validar_seleccion_combos(diccionario_campos_oblig_combos)==False:
            return
        
        wb_req=openpyxl.load_workbook(path_db)
        wb_listas=openpyxl.load_workbook(path_config)
        
        
        hist_req_sheet=wb_req['Requerimientos']
        req2_sheet=wb_listas['Requerimientos-12']

        self.fila_vacia = 1
        
        while (hist_req_sheet.cell(row = self.fila_vacia, column = 1).value != None) :
          self.fila_vacia += 1
        
        for cell in hist_req_sheet['A']:
                if cell.value !=None:
                    self.lista_nro_req.append(cell.value)
        try:
            self.nro_req=int(self.lista_nro_req[-1])+1
        except:
            self.nro_req=1

        requerimiento_auto=self.nro_req
        fecha_auto=self.lblfecha_auto.GetLabel()
        cotizacion=self.txtcotizacion.GetValue()
        nombre_cliente=self.txtnombrecliente.GetValue()
        tiporeq=self.combotiporeq.GetValue()
        tipotransporte=self.combotipotransporte.GetValue()
        tipocontenedor=self.combotipocontenedor.GetValue()
        requieredescargue=self.comborequieredescargue.GetValue()
        origen=self.txtorigen.GetValue()
        destino=self.txtdestino.GetValue()
        km=self.txtkm.GetValue()
        precio=self.txtprecio.GetValue()
        nombreresponsable=self.txtnombreresponsable.GetValue()
        telefono_resp=self.txttelefono_resp.GetValue()
        cargo=self.txtcargo.GetValue()
        nombresiso=self.txtnombresiso.GetValue()
        telefono_siso=self.txttelefono_siso.GetValue()
        
        horasantes=self.txthorasantes.GetValue()
        
        if self.check_si_peaje.IsChecked():
            check_peaje="Si"
        else:
            check_peaje="No"
            
        if self.check_si_info.IsChecked():
            debeinfo="Si"
        else:
            debeinfo="No"
         
        self.dic_asosiacion={}
        self.lista_asociacion=[]
        self.lista_tipo_transp2=[]
        
        for cell in req2_sheet['D']:
            if cell != None:
                self.lista_asociacion.append(cell.value)
        
        for cell in req2_sheet['C']:
            if cell != None:
                self.lista_tipo_transp2.append(cell.value)
        
        
        for i in range((len(self.lista_tipo_transp2))):
            self.dic_asosiacion[self.lista_tipo_transp2[i]]=self.lista_asociacion[i]
        
        
        
        hist_req_sheet.cell(row=self.fila_vacia, column=col_requerimiento_auto).value=requerimiento_auto
        hist_req_sheet.cell(row=self.fila_vacia, column=col_area_req).value=self.area_selec
        hist_req_sheet.cell(row=self.fila_vacia, column=col_area).value=self.dic_asosiacion[tipotransporte]
        hist_req_sheet.cell(row=self.fila_vacia, column=col_fecha_auto).value=fecha_auto
        hist_req_sheet.cell(row=self.fila_vacia, column=col_cotizacion).value=cotizacion
        hist_req_sheet.cell(row=self.fila_vacia, column=col_nombrecliente).value=nombre_cliente
        hist_req_sheet.cell(row=self.fila_vacia, column=col_tipo_req).value=tiporeq
        hist_req_sheet.cell(row=self.fila_vacia, column=col_tipotransporte).value=tipotransporte
        hist_req_sheet.cell(row=self.fila_vacia, column=col_tipocontenedor).value=tipocontenedor
        hist_req_sheet.cell(row=self.fila_vacia, column=col_requieredescargue).value=requieredescargue
        hist_req_sheet.cell(row=self.fila_vacia, column=col_origen).value=origen
        hist_req_sheet.cell(row=self.fila_vacia, column=col_destino).value=destino
        hist_req_sheet.cell(row=self.fila_vacia, column=col_km).value=km
        hist_req_sheet.cell(row=self.fila_vacia, column=col_precio).value=precio
        hist_req_sheet.cell(row=self.fila_vacia, column=col_nombreresponsable).value=nombreresponsable
        hist_req_sheet.cell(row=self.fila_vacia, column=col_telefono_resp).value=telefono_resp
        hist_req_sheet.cell(row=self.fila_vacia, column=col_cargo).value=cargo
        hist_req_sheet.cell(row=self.fila_vacia, column=col_nombresiso).value=nombresiso
        hist_req_sheet.cell(row=self.fila_vacia, column=col_telefono_siso).value=telefono_siso
        hist_req_sheet.cell(row=self.fila_vacia, column=col_debeinfo).value=debeinfo
        hist_req_sheet.cell(row=self.fila_vacia, column=col_horasantes).value=horasantes
        hist_req_sheet.cell(row=self.fila_vacia, column=col_recargopeaje).value=check_peaje
        self.precio_final(hist_req_sheet)
        
        
        
        try:
            wb_req.save(path_db)
            
            self.precio_final(hist_req_sheet)
            self.txtcotizacion.Value=''
            self.combotipotransporte.Value=self.lista_tipo_transp[0]
            self.combotipocontenedor.Value=self.lista_tipo_cont[0]
            self.comborequieredescargue.Value=self.lista_descargue[0]
            self.combotiporeq.Value=self.lista_tiporeq[0]
            self.txtorigen.Value=''
            self.txtnombrecliente.Value=''
            self.txtdestino.Value=''
            self.txtkm.Value=''
            self.txtprecio.Value=''
            self.txtnombreresponsable.Value=''
            self.txttelefono_resp.Value=''
            self.txtcargo.Value=''
            self.txtnombresiso.Value=''
            self.txttelefono_siso.Value=''
            self.txthorasantes.Value=''
            self.check_no_peaje.SetValue(True)
            self.check_si_peaje.SetValue(False)
            self.check_no_info.SetValue(False)
            self.check_si_info.SetValue(False)
            
              
            
            sheet_config=wb_listas['Config']
            area_correo=self.dic_asosiacion[tipotransporte]
            
            if area_correo=='Operaciones':
                receiver=sheet_config.cell(row=5, column=2).value
                
            elif area_correo=='Administracion':
                receiver=sheet_config.cell(row=6,column=2).value
            else:
                receiver=sheet_config.cell(row=7,column=2).value

            self.enviar_email(receiver, self.nro_req, area_correo)
            
            
            for cell in hist_req_sheet['A']:
                if cell.value !=None:
                    self.lista_nro_req.append(cell.value)
            self.nro_req= int(self.lista_nro_req[-1])+1 
            self.Destroy()
            
             
        except Exception as e:
            print(e)
            error_msgbox=wx.MessageDialog(None,'Error al guardar el registro en la BD. \nVerifique el el archivo de excel este cerrado y en la ruta correcta.','ERROR',wx.ICON_ERROR)
            error_msgbox.ShowModal()
        
    
    def enviar_email(self, receiver, nro_req, area_encargada):
        
        EMAIL_ADDRESS='requerimientologisticocontant@gmail.com'
        EMAIL_PASSWORD='pewljcgvqnrjhegz'
        
        msg = EmailMessage()
        msg['Subject'] = 'Nuevo Requerimiento Logistico No. ' + str(nro_req) + " // " + str(datetime.today().strftime('%d-%m-%Y'))
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = receiver
    
    
        initial_html="""\
        <!DOCTYPE html>
        <html>
        <head>
        <style>
        html{
            font-family: Arial, Helvetica, sans-serif;
        }
        h1{
            color:green;
        }
        table{
            border-collapse: collapse;
        }
        th,td{
            border: 1px solid black;
        }
        
        </style>
        </head>
            <body>
                <h1>Nuevo Requerimiento</h1>
                <p>Usted Tiene un Nuevo Requerimiento No : """+str(nro_req)+""" </p>
                <p>Area Encargada : """+area_encargada+""" </p>
                <p>Favor diríjase al centro logistico para darle tramite al requerimiento </p>
                <p>Contenedores de Antioquia</p>    
            </body>
        </html>"""
        
    
        msg.add_alternative(initial_html, subtype='html')
    
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)


    def salir(self,event):
        salir_msgbox=wx.MessageBox('¿Esta seguro que desea salir sin guardar?','Salir sin Guardar',wx.YES_NO| wx.ICON_WARNING)
        
        if salir_msgbox == wx.YES:
            self.Destroy()
        else:
            pass
        
    
    def adicionar_transp(self,event):
        
        if self.check_si_info.IsChecked():
            diccionario_campos_oblig_texto={self.txtcotizacion:'Cotizacion N°',  self.txtnombrecliente:'Nombre Cliente',self.txtorigen:'Origen', self.txtdestino:'Destino', self.txtnombreresponsable:'Nombre Responsable', self.txttelefono_resp:'Telefono Resposnable', self.txthorasantes:'N° Horas Antes'}
        else:
            diccionario_campos_oblig_texto={self.txtcotizacion:'Cotizacion N°',  self.txtnombrecliente:'Nombre Cliente',self.txtorigen:'Origen', self.txtdestino:'Destino', self.txtnombreresponsable:'Nombre Responsable', self.txttelefono_resp:'Telefono Resposnable'}
        
        diccionario_campos_oblig_numero={self.txtprecio:'Precio'}
        diccionario_campos_oblig_combos={self.combotipotransporte:'Tipo de Transporte', self.combotiporeq:'Tipo Requerimiento', self.combotipocontenedor:'Tipo de Contenedor', self.comborequieredescargue:'Requiere Desacargue'} 
        
        if self.validar_campos_vacios_texto(diccionario_campos_oblig_texto)==False or self.validar_campos_vacios_numero(diccionario_campos_oblig_numero)==False or self.validar_seleccion_combos(diccionario_campos_oblig_combos)==False:
            return
           
        wb_req=openpyxl.load_workbook(path_db)
        wb_listas=openpyxl.load_workbook(path_config)
        
        hist_req_sheet=wb_req['Requerimientos']
        req2_sheet=wb_listas['Requerimientos-12']
        self.fila_vacia = 1
        
        while (hist_req_sheet.cell(row = self.fila_vacia, column = 1).value != None) :
          self.fila_vacia += 1
        
        for cell in hist_req_sheet['A']:
                if cell.value !=None:
                    self.lista_nro_req.append(cell.value)
        try:
            self.nro_req=int(self.lista_nro_req[-1])+1
        except:
            self.nro_req=1
        
        requerimiento_auto=self.nro_req
        fecha_auto=self.lblfecha_auto.GetLabel()
        cotizacion=self.txtcotizacion.GetValue()
        nombre_cliente=self.txtnombrecliente.GetValue()
        tiporeq=self.combotiporeq.GetValue()
        tipotransporte=self.combotipotransporte.GetValue()
        tipocontenedor=self.combotipocontenedor.GetValue()
        requieredescargue=self.comborequieredescargue.GetValue()
        origen=self.txtorigen.GetValue()
        destino=self.txtdestino.GetValue()
        km=self.txtkm.GetValue()
        precio=self.txtprecio.GetValue()
        nombreresponsable=self.txtnombreresponsable.GetValue()
        telefono_resp=self.txttelefono_resp.GetValue()
        cargo=self.txtcargo.GetValue()
        nombresiso=self.txtnombresiso.GetValue()
        telefono_siso=self.txttelefono_siso.GetValue()
        
        horasantes=self.txthorasantes.GetValue()
        
        if self.check_si_peaje.IsChecked():
            check_peaje="Si"
        else:
            check_peaje="No"
       
        if self.check_si_info.IsChecked():
            debeinfo="Si"
        else:
            debeinfo="No"
        
        self.dic_asosiacion={}
        self.lista_asociacion=[]
        self.lista_tipo_transp2=[]
        
        for cell in req2_sheet['D']:
            if cell != None:
                self.lista_asociacion.append(cell.value)
        
        for cell in req2_sheet['C']:
            if cell != None:
                self.lista_tipo_transp2.append(cell.value)
        
        
        for i in range((len(self.lista_tipo_transp2))):
            self.dic_asosiacion[self.lista_tipo_transp2[i]]=self.lista_asociacion[i]
        
        hist_req_sheet.cell(row=self.fila_vacia, column=col_requerimiento_auto).value=requerimiento_auto
        hist_req_sheet.cell(row=self.fila_vacia, column=col_area_req).value=self.area_selec
        hist_req_sheet.cell(row=self.fila_vacia, column=col_area).value=self.dic_asosiacion[tipotransporte]
        hist_req_sheet.cell(row=self.fila_vacia, column=col_fecha_auto).value=fecha_auto
        hist_req_sheet.cell(row=self.fila_vacia, column=col_cotizacion).value=cotizacion
        hist_req_sheet.cell(row=self.fila_vacia, column=col_nombrecliente).value=nombre_cliente
        hist_req_sheet.cell(row=self.fila_vacia, column=col_tipo_req).value=tiporeq
        hist_req_sheet.cell(row=self.fila_vacia, column=col_tipotransporte).value=tipotransporte
        hist_req_sheet.cell(row=self.fila_vacia, column=col_tipocontenedor).value=tipocontenedor
        hist_req_sheet.cell(row=self.fila_vacia, column=col_requieredescargue).value=requieredescargue
        hist_req_sheet.cell(row=self.fila_vacia, column=col_origen).value=origen
        hist_req_sheet.cell(row=self.fila_vacia, column=col_destino).value=destino
        hist_req_sheet.cell(row=self.fila_vacia, column=col_km).value=km
        hist_req_sheet.cell(row=self.fila_vacia, column=col_precio).value=precio
        hist_req_sheet.cell(row=self.fila_vacia, column=col_nombreresponsable).value=nombreresponsable
        hist_req_sheet.cell(row=self.fila_vacia, column=col_telefono_resp).value=telefono_resp
        hist_req_sheet.cell(row=self.fila_vacia, column=col_cargo).value=cargo
        hist_req_sheet.cell(row=self.fila_vacia, column=col_nombresiso).value=nombresiso
        hist_req_sheet.cell(row=self.fila_vacia, column=col_telefono_siso).value=telefono_siso
        hist_req_sheet.cell(row=self.fila_vacia, column=col_debeinfo).value=debeinfo
        hist_req_sheet.cell(row=self.fila_vacia, column=col_horasantes).value=horasantes
        hist_req_sheet.cell(row=self.fila_vacia, column=col_recargopeaje).value=check_peaje
        self.precio_final(hist_req_sheet)
        
        self.combotipotransporte.Value=self.lista_tipo_transp[0]
        self.combotipocontenedor.Value=self.lista_tipo_cont[0]
        self.txtprecio.Value=''
        self.comborequieredescargue.Value=self.lista_tipo_transp[0]
        
        try:
            wb_req.save(path_db)
            self.lista_nro_req=[]
            for cell in hist_req_sheet['A']:
                if cell.value !=None:
                    self.lista_nro_req.append(cell.value)
            
            
            sheet_config=wb_listas['Config']
            area_correo=self.dic_asosiacion[tipotransporte]
            
            if area_correo=='Operaciones':
                receiver=sheet_config.cell(row=5, column=2).value
                
            elif area_correo=='Administracion':
                receiver=sheet_config.cell(row=6,column=2).value
            else:
                receiver=sheet_config.cell(row=7,column=2).value

            self.enviar_email(receiver, self.nro_req, area_correo)
            
            
            self.nro_req= int(self.lista_nro_req[-1])+1 
            self.lblrequerimiento.SetLabel(label='Requerimiento N° ' + str(self.nro_req))
        except Exception as e:
            print(e)
            error_msgbox=wx.MessageDialog(None,'Error al guardar el registro en la BD. \nVerifique el el archivo de excel este cerrado y en la ruta correcta.','ERROR',wx.ICON_ERROR)
            error_msgbox.ShowModal()        
            
        
    
    def validar_campos_vacios_texto(self,diccionario_campos_oblig):
        for campo in diccionario_campos_oblig:
            if len(campo.GetValue().strip()) == 0:
                error_msgbox=wx.MessageDialog(None,'Falta diligenciar el campo: ' + diccionario_campos_oblig[campo],'ERROR',wx.ICON_ERROR)
                error_msgbox.ShowModal()   
                return False
        return True
    
    def validar_campos_vacios_numero(self, diccionario_campos_oblig):
        
        for campo in diccionario_campos_oblig:
            if campo.GetValue().isnumeric() == False:
                error_msgbox=wx.MessageDialog(None,'El campo: ' + diccionario_campos_oblig[campo] +' Solo debe contener caracteres numericos','ERROR',wx.ICON_ERROR)
                error_msgbox.ShowModal()   
                return False
        return True
    
    def validar_seleccion_combos(self,diccionario_campos_oblig):
        
        for campo in diccionario_campos_oblig:
            if campo.GetSelection()== 0 or campo.GetSelection()== -1:
                error_msgbox=wx.MessageDialog(None,'Seleccione una opcion en el campo ' + diccionario_campos_oblig[campo],'ERROR',wx.ICON_ERROR)
                error_msgbox.ShowModal()   
                return False
        return True

class NuevoReqPanel(wx.Panel):

    def __init__(self,parent):
        # create the panel
        wx.Panel.__init__(self, parent=parent)
        try:

            image_file = 'logo35.png'
            bmp1 = wx.Image(
                image_file, 
                wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            # image's upper left corner anchors at panel 
            # coordinates (0, 0)
            self.bitmap2 = wx.StaticBitmap(
                self, -1, bmp1, (5, 0))
            # show some image details
            #str1 = "%s  %dx%d" % (image_file, bmp1.GetWidth(),
                                  #bmp1.GetHeight()) 
            #parent.SetTitle(str1)
        except IOError:
            print ("Image file %s not found")
            raise SystemExit        


class ww_logistica21(wx.Frame):
    
    def __init__(self,parent):
        
        wb_listas=openpyxl.load_workbook(path_config)
        wb_req=openpyxl.load_workbook(path_db)

        wx.Frame.__init__(self, None, wx.ID_ANY, "Centro Logistico", size=(270, 250),style=wx.DEFAULT_FRAME_STYLE & ~(wx.RESIZE_BORDER | wx.MAXIMIZE_BOX))  
        self.SetBackgroundColour(secondary_color)
        self.Center()
        try:
            
            #image_file = 'CINCO CONSULTORES.jpg'
            #bmp1 = wx.Image(
                #image_file, 
                #wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            #self.panel = wx.StaticBitmap(
                #self, -1, bmp1, (0, 0)
            self.panel=wx.Panel(self)
            panel_font= wx.Font(10, wx.DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Normal")
            self.panel.SetBackgroundColour(secondary_color)

        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
        
        ico = wx.Icon('Cont.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        self.fgs= wx.GridBagSizer(0,0)
        
        title_font= wx.Font(11, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Normal")
       
        self.lbltitle =wx.StaticText(self.panel, label='Ingrese Numero de Requerimiento\n a Gestionar:')
        self.lbltitle.SetFont(title_font)
        self.lbltitle.SetBackgroundColour(secondary_color)
        self.lbltitle.SetForegroundColour(principal_color)
        self.fgs.Add(self.lbltitle,pos=(2,1),span=(1,3), flag=wx.ALL | wx.ALIGN_CENTER, border=5)

        self.txtreq = wx.TextCtrl(self.panel)
        self.fgs.Add(self.txtreq , pos=(4,1),span=(1,3), flag= wx.ALL| wx.ALIGN_CENTER, border=5)
        
        btn_aceptar = wx.Button(self.panel, id=wx.ID_OK, label="Aceptar",size=(-1,-1))
        self.fgs.Add(btn_aceptar, pos=(6,1),span=(1,3), flag= wx.ALL | wx.ALIGN_CENTER, border=0)
        btn_aceptar.Bind(wx.EVT_BUTTON, self.open_logistica22)

        mainSizer= wx.BoxSizer(wx.VERTICAL)
        mainSizer.Add(self.fgs,0, flag=wx.ALIGN_LEFT)
        self.panel.SetSizerAndFit(mainSizer)
        
    #-------------Button Functions-----------------#            
    def open_logistica22(self, event):
        
        wb_req=openpyxl.load_workbook(path_db)


        
        hist_req_sheet=wb_req['Requerimientos']
        self.lista_nro_req=[]
        
        for cell in hist_req_sheet['A']:
            if cell.value !=None:
                self.lista_nro_req.append(cell.value)
        global req_selec
        try:
            req_selec=int(self.txtreq.GetValue())
        except:
            error_msgbox=wx.MessageDialog(None,'Numero de Requerimiento No Encontrado','ERROR',wx.ICON_ERROR)
            error_msgbox.ShowModal()
            return
                
        if req_selec in self.lista_nro_req:
            ww_logistica22(parent=self.panel).Show() 
            self.Destroy()
        else:
            error_msgbox=wx.MessageDialog(None,'Numero de Requerimiento No Encontrado','ERROR',wx.ICON_ERROR)
            error_msgbox.ShowModal()    
           
class ww_logistica22(wx.Frame):
    def __init__(self,parent):    
         
        wb_listas=openpyxl.load_workbook(path_config)
        
        
        req2_sheet=wb_listas['Requerimientos-12']
        wb_req=openpyxl.load_workbook(path_db)
        self.hist_req_sheet=wb_req['Requerimientos']
        
        self.lista_descargue=[]
        for cell in req2_sheet['E']:
            if cell.value != None:
                self.lista_descargue.append(cell.value)

        
        global req_selec
        
        self.lista_requerimientos=[]
        
        for cell in self.hist_req_sheet['A']:
            if cell.value != None:
                self.lista_requerimientos.append(cell.value)

        self.nro_fila_req=int(self.lista_requerimientos.index(req_selec))+1
        
        #make a list thtat contains every data in the row
        self.lista_valores_fila=[]
        for cell in self.hist_req_sheet[self.nro_fila_req]:
            self.lista_valores_fila.append(cell.value)
        
        #----------Front------------#
        wx.Frame.__init__(self, None, wx.ID_ANY, "Centro Logistico", size=(1020, 670),style=wx.DEFAULT_FRAME_STYLE & ~(wx.RESIZE_BORDER | wx.MAXIMIZE_BOX))
        self.panel = LogisticaPanel(self)
        panel_font= wx.Font(10, wx.DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Normal")
        self.panel.SetFont(panel_font)
        self.SetBackgroundColour(secondary_color)
        self.Center()
        
        ico = wx.Icon('Cont.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        self.fgs= wx.GridBagSizer(0,0)
        
        title_font= wx.Font(25, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Bold")
        title_font3= wx.Font(15, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Bold")
        bold_font= wx.Font(10, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Bold")
    
        
        self.lbltitle=wx.StaticText(self.panel, label='LOGISTICA')
        self.lblrequerimiento=wx.StaticText(self.panel, label='Requerimiento N°  ' + str(self.lista_valores_fila[col_requerimiento_auto-1]))
        self.lblfecha=wx.StaticText(self.panel, label='Fecha  '+ self.lista_valores_fila[col_fecha_auto-1])
        self.lblareaencargada=wx.StaticText(self.panel, label='Area Encargada:  ' + self.lista_valores_fila[col_area-1])
        self.lblcotizacion=wx.StaticText(self.panel, label='Cotizacion N°')
        self.lbltiporeq=wx.StaticText(self.panel, label='Tipo Requerimiento')
        self.lbltipotransp=wx.StaticText(self.panel, label='Tipo de Transporte')
        self.lbltipocont=wx.StaticText(self.panel, label='Tipo de Contenedor')
        self.lbldescargue=wx.StaticText(self.panel, label='Requiere Descargue')
        self.lblorigen=wx.StaticText(self.panel, label='Origen')
        self.lbldestino=wx.StaticText(self.panel, label='Destino')
        self.lblkm=wx.StaticText(self.panel, label='Km')
        self.lblprecio=wx.StaticText(self.panel, label='Precio')
        self.lblrecargopeaje=wx.StaticText(self.panel, label='Recargo Peaje')
        self.lblnombreresp=wx.StaticText(self.panel, label='Nombre\nResponsable')
        self.lbltelresp=wx.StaticText(self.panel, label='Telefono Resp.')
        self.lblcargoresp=wx.StaticText(self.panel, label='Cargo')
        self.lblnombresiso=wx.StaticText(self.panel, label='Nombre SISO')
        self.lbltelesiso=wx.StaticText(self.panel, label='Telefono SISO')
        self.lbldebeinfo=wx.StaticText(self.panel, label='Debe Enviarse\nInformacion')
        self.lblhorasantes=wx.StaticText(self.panel, label='N° de Horas Antes')
        self.lblinfologistica=wx.StaticText(self.panel, label='Info Logistica')
        self.lblinfocliente=wx.StaticText(self.panel, label='Info Cliente')
        self.lblfechaentrega=wx.StaticText(self.panel, label='Fecha de Entrega')
        self.lbldireccion=wx.StaticText(self.panel, label='Direccion Exacta')
        self.lblreferenciacont=wx.StaticText(self.panel, label='Ref.Contenedor')
        self.lblnombreconduc=wx.StaticText(self.panel, label='Nombre Conductor')
        self.lblcedula=wx.StaticText(self.panel, label='Cedula')
        self.lbltelefonoconduc=wx.StaticText(self.panel, label='Telefono')
        self.lblplaca=wx.StaticText(self.panel, label='Placa')
        self.lbladiciones=wx.StaticText(self.panel, label='Adiciones Entrega')
        
        if self.lista_valores_fila[col_horasantes-1] != None:
            horas=str(self.lista_valores_fila[col_horasantes-1])
        else:
            horas='0'
            
        self.lblpreguntahoras=wx.StaticText(self.panel, label='Documentacion Enviada '+ horas + ' Horas Antes?')
        self.lblpreguntadoc=wx.StaticText(self.panel, label='Documentacion Completa?')
        self.lblnombrecliente=wx.StaticText(self.panel, label='Nombre Cliente')

        self.txtcotizacion=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txttiporeq=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txttipotransp=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txttipocont=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txtdescargue=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txtorigen=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txtdestino=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txtkm=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txtprecio=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txtrecargopeaje=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txtnombreresp=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txttelresp=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txtcargoresp=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txtnombresiso=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txttelesiso=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txtdebeinfo=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txthorasantes=wx.TextCtrl(self.panel,style=wx.TE_READONLY)
        self.txtnombrecliente=wx.TextCtrl(self.panel,style=wx.TE_READONLY)

        try:
            self.txtcotizacion.SetValue(self.lista_valores_fila[col_cotizacion-1])
        except:
            self.txtcotizacion.SetValue('N/A')
    
        try:
            self.txttiporeq.SetValue(self.lista_valores_fila[col_tipo_req-1])
        except:
            self.txttiporeq.SetValue('N/A')
        try:
            self.txttipotransp.SetValue(self.lista_valores_fila[col_tipotransporte-1])
        except:
            self.txttipotransp.SetValue('N/A')
        try:
            self.txttipocont.SetValue(self.lista_valores_fila[col_tipocontenedor-1])
        except:
            self.txttipocont.SetValue('N/A')
    
        try:
            self.txtdescargue.SetValue(self.lista_valores_fila[col_requieredescargue-1])
        except:
            self.txtdescargue.SetValue('N/A')  
        try:
            self.txtorigen.SetValue(self.lista_valores_fila[col_origen-1])
        except:
            self.txtorigen.SetValue('N/A')
        try:
            self.txtdestino.SetValue(self.lista_valores_fila[col_destino-1])
        except:
            self.txtdestino.SetValue('N/A')
        try:
            self.txtkm.SetValue(self.lista_valores_fila[col_km-1])
        except:
            self.txtkm.SetValue('N/A')
        try:
            self.txtprecio.SetValue(str(self.lista_valores_fila[col_precio_recargo-1]))
        except:

            self.txtprecio.SetValue('N/A')
        try:
            self.txtrecargopeaje.SetValue(self.lista_valores_fila[col_recargopeaje-1])
        except:
            self.txtrecargopeaje.SetValue('N/A')
        try:
            self.txtnombreresp.SetValue(self.lista_valores_fila[col_nombreresponsable-1])
        except:
            self.txtnombreresp.SetValue('N/A')
        try:
            self.txttelresp.SetValue(self.lista_valores_fila[col_telefono_resp-1])
        except:
            self.txttelresp.SetValue('N/A')
        try:
            self.txtcargoresp.SetValue(self.lista_valores_fila[col_cargo-1])
        except:
            self.txtcargoresp.SetValue('N/A')
        try:
            self.txtnombresiso.SetValue(self.lista_valores_fila[col_nombresiso-1])
        except:
            self.txtnombresiso.SetValue('N/A')
        try:
            self.txttelesiso.SetValue(self.lista_valores_fila[col_telefono_siso-1])
        except:
            self.txttelesiso.SetValue('N/A')
        try:
            self.txtdebeinfo.SetValue(self.lista_valores_fila[col_debeinfo-1])
        except:
            self.txtdebeinfo.SetValue('N/A')
        try:
            self.txthorasantes.SetValue(str(self.lista_valores_fila[col_horasantes-1]))
        except:
            self.txthorasantes.SetValue('N/A')
            
        try:
            self.txtnombrecliente.SetValue(str(self.lista_valores_fila[col_nombrecliente-1]))
        except:
            self.txtnombrecliente.SetValue('N/A')


        
        
        self.txtfechaentrega=wx.TextCtrl(self.panel)
        self.txtdireccion=wx.TextCtrl(self.panel)
        self.txtreferenciacont=wx.TextCtrl(self.panel)
        self.txtplaca=wx.TextCtrl(self.panel)
        self.txttelefonoconduc=wx.TextCtrl(self.panel)
        self.txtcedula=wx.TextCtrl(self.panel)
        self.txtnombreconduc=wx.TextCtrl(self.panel)
        self.txtadiciones=wx.TextCtrl(self.panel,style = wx.TE_MULTILINE)
        self.checkpreguntadoc_si=wx.CheckBox(self.panel, label= "Si")
        self.checkpreguntadoc_no=wx.CheckBox(self.panel, label= "No")
        self.checkpreguntahoras_si=wx.CheckBox(self.panel, label= "Si")
        self.checkpreguntahoras_no=wx.CheckBox(self.panel, label= "No")
        
        
        try:
            self.txtfechaentrega.SetValue(self.lista_valores_fila[col_fechaentrega-1])
        except:
            self.txtfechaentrega.SetValue('')
        
        try:
            self.txtdireccion.SetValue(self.lista_valores_fila[col_direccion-1])
        except:
            self.txtdireccion.SetValue('')
        
        try:
            self.txtreferenciacont.SetValue(self.lista_valores_fila[col_referenciacont-1])
        except:
            self.txtreferenciacont.SetValue('')
            
        try:
            self.txtplaca.SetValue(self.lista_valores_fila[col_placa-1])
        except:
            self.txtplaca.SetValue('')
            
        try:
            self.txttelefonoconduc.SetValue(self.lista_valores_fila[col_telefonoconduc-1])
        except:
            self.txttelefonoconduc.SetValue('')
        
        try:
            self.txtcedula.SetValue(self.lista_valores_fila[col_cedula-1])
        except:
            self.txtcedula.SetValue('')
        
        try:
            self.txtnombreconduc.SetValue(self.lista_valores_fila[col_nombreconduc-1])
        except:
            self.txtnombreconduc.SetValue('')
        
        try:
            self.txtadiciones.SetValue(self.lista_valores_fila[col_adiciones-1])
        except:
            self.txtadiciones.SetValue('')

        
        
        
        self.lblrequerimiento.SetFont(bold_font)
        self.lblfecha.SetFont(bold_font)
        self.lblareaencargada.SetFont(bold_font)
        self.lblcotizacion.SetFont(bold_font)
        self.lbltipotransp.SetFont(bold_font)
        self.lbltipocont.SetFont(bold_font)
        self.lbldescargue.SetFont(bold_font)
        self.lblorigen.SetFont(bold_font)
        self.lbldestino.SetFont(bold_font)
        self.lblkm.SetFont(bold_font)
        self.lblprecio.SetFont(bold_font)
        self.lblrecargopeaje.SetFont(bold_font)
        self.lblnombreresp.SetFont(bold_font)
        self.lbltelresp.SetFont(bold_font)
        self.lblcargoresp.SetFont(bold_font)
        self.lblnombresiso.SetFont(bold_font)
        self.lbltelesiso.SetFont(bold_font)
        self.lbldebeinfo.SetFont(bold_font)
        self.lblhorasantes.SetFont(bold_font)
        self.lblinfologistica.SetFont(title_font3)
        self.lblinfocliente.SetFont(title_font3)
        self.lblfechaentrega.SetFont(bold_font)
        self.lbldireccion.SetFont(bold_font)
        self.lblreferenciacont.SetFont(bold_font)
        self.lblnombreconduc.SetFont(bold_font)
        self.lblcedula.SetFont(bold_font)
        self.lbltelefonoconduc.SetFont(bold_font)
        self.lblplaca.SetFont(bold_font)
        self.lbladiciones.SetFont(bold_font)
        self.lblpreguntahoras.SetFont(bold_font)
        self.lblpreguntadoc.SetFont(bold_font)
        self.lbltiporeq.SetFont(bold_font)
        self.lblnombrecliente.SetFont(bold_font)

        self.lbltitle.SetBackgroundColour(secondary_color)
        self.lblrequerimiento.SetBackgroundColour(secondary_color)
        self.lblfecha.SetBackgroundColour(secondary_color)
        self.lblareaencargada.SetBackgroundColour(secondary_color)
        self.lblcotizacion.SetBackgroundColour(secondary_color)
        self.lbltipotransp.SetBackgroundColour(secondary_color)
        self.lbltipocont.SetBackgroundColour(secondary_color)
        self.lbldescargue.SetBackgroundColour(secondary_color)
        self.lblorigen.SetBackgroundColour(secondary_color)
        self.lbldestino.SetBackgroundColour(secondary_color)
        self.lblkm.SetBackgroundColour(secondary_color)
        self.lblprecio.SetBackgroundColour(secondary_color)
        self.lblrecargopeaje.SetBackgroundColour(secondary_color)
        self.lblnombreresp.SetBackgroundColour(secondary_color)
        self.lbltelresp.SetBackgroundColour(secondary_color)
        self.lblcargoresp.SetBackgroundColour(secondary_color)
        self.lblnombresiso.SetBackgroundColour(secondary_color)
        self.lbltelesiso.SetBackgroundColour(secondary_color)
        self.lbldebeinfo.SetBackgroundColour(secondary_color)
        self.lblhorasantes.SetBackgroundColour(secondary_color)
        self.lblinfologistica.SetBackgroundColour(secondary_color)
        self.lblinfocliente.SetBackgroundColour(secondary_color)
        self.lblfechaentrega.SetBackgroundColour(secondary_color)
        self.lbldireccion.SetBackgroundColour(secondary_color)
        self.lblreferenciacont.SetBackgroundColour(secondary_color)
        self.lblnombreconduc.SetBackgroundColour(secondary_color)
        self.lblcedula.SetBackgroundColour(secondary_color)
        self.lbltelefonoconduc.SetBackgroundColour(secondary_color)
        self.lblplaca.SetBackgroundColour(secondary_color)
        self.lbladiciones.SetBackgroundColour(secondary_color)
        self.lblpreguntahoras.SetBackgroundColour(secondary_color)
        self.lblpreguntadoc.SetBackgroundColour(secondary_color)
        self.lbltiporeq.SetBackgroundColour(secondary_color)

        self.checkpreguntahoras_si.SetBackgroundColour(secondary_color)
        self.checkpreguntahoras_no.SetBackgroundColour(secondary_color)
        self.checkpreguntadoc_si.SetBackgroundColour(secondary_color)
        self.checkpreguntadoc_no.SetBackgroundColour(secondary_color)

        self.lbltitle.SetForegroundColour(principal_color)
        self.lblrequerimiento.SetForegroundColour(principal_color)
        self.lblfecha.SetForegroundColour(principal_color)
        self.lblareaencargada.SetForegroundColour(principal_color)
        self.lblcotizacion.SetForegroundColour(principal_color)
        self.lbltipotransp.SetForegroundColour(principal_color)
        self.lbltipocont.SetForegroundColour(principal_color)
        self.lbldescargue.SetForegroundColour(principal_color)
        self.lblorigen.SetForegroundColour(principal_color)
        self.lbldestino.SetForegroundColour(principal_color)
        self.lblkm.SetForegroundColour(principal_color)
        self.lblprecio.SetForegroundColour(principal_color)
        self.lblrecargopeaje.SetForegroundColour(principal_color)
        self.lblnombreresp.SetForegroundColour(principal_color)
        self.lbltelresp.SetForegroundColour(principal_color)
        self.lblcargoresp.SetForegroundColour(principal_color)
        self.lblnombresiso.SetForegroundColour(principal_color)
        self.lbltelesiso.SetForegroundColour(principal_color)
        self.lbldebeinfo.SetForegroundColour(principal_color)
        self.lblhorasantes.SetForegroundColour(principal_color)
        self.lblinfologistica.SetForegroundColour(principal_color)
        self.lblinfocliente.SetForegroundColour(principal_color)
        self.lblfechaentrega.SetForegroundColour(principal_color)
        self.lbldireccion.SetForegroundColour(principal_color)
        self.lblreferenciacont.SetForegroundColour(principal_color)
        self.lblnombreconduc.SetForegroundColour(principal_color)
        self.lblcedula.SetForegroundColour(principal_color)
        self.lbltelefonoconduc.SetForegroundColour(principal_color)
        self.lblplaca.SetForegroundColour(principal_color)
        self.lbladiciones.SetForegroundColour(principal_color)
        self.lblpreguntahoras.SetForegroundColour(principal_color)
        self.lblpreguntadoc.SetForegroundColour(principal_color)
        self.lbltiporeq.SetForegroundColour(principal_color)
        self.lblnombrecliente.SetForegroundColour(principal_color)

        self.checkpreguntahoras_si.SetForegroundColour(principal_color)
        self.checkpreguntahoras_no.SetForegroundColour(principal_color)
        self.checkpreguntadoc_si.SetForegroundColour(principal_color)
        self.checkpreguntadoc_no.SetForegroundColour(principal_color)
        
        
        btn_finalizar = wx.Button(self.panel, id=wx.ID_ANY, label="Finalizar",size=(-1,-1))
        
        self.lbltitle.SetFont(title_font)
        self.lblrequerimiento.SetFont(title_font3)

        
        self.fgs.Add(self.lbltitle,pos=(1,1),span=(1,8), flag= wx.ALL  | wx.ALIGN_CENTER, border=5)
        self.fgs.Add(self.lblrequerimiento,pos=(2,1),span=(1,2), flag= wx.ALL, border=0)
        self.fgs.Add(self.lblfecha,pos=(3,1),span=(1,1), flag= wx.ALL, border=0)
        self.fgs.Add(self.lblareaencargada,pos=(4,1),span=(1,2), flag= wx.ALL, border=0)
        self.fgs.Add(self.lblcotizacion,pos=(7,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbltiporeq,pos=(8,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbltipotransp,pos=(9,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbltipocont,pos=(10,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbldescargue,pos=(11,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblorigen,pos=(7,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbldestino,pos=(8,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblkm,pos=(9,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblprecio,pos=(10,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblrecargopeaje,pos=(11,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblnombreresp,pos=(7,5),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbltelresp,pos=(8,5),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblcargoresp,pos=(9,5),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblnombresiso,pos=(10,5),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbltelesiso,pos=(9,7),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbldebeinfo,pos=(7,7),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblhorasantes,pos=(8,7),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblinfologistica,pos=(6,1),span=(1,4), flag= wx.ALL | wx.ALIGN_CENTER, border=5)
        self.fgs.Add(self.lblinfocliente,pos=(6,5),span=(1,4), flag= wx.ALL| wx.ALIGN_CENTER, border=5)
        self.fgs.Add(self.lblfechaentrega,pos=(13,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbldireccion,pos=(13,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblreferenciacont,pos=(13,7),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblnombreconduc,pos=(14,4),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblcedula,pos=(14,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbltelefonoconduc,pos=(14,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lblplaca,pos=(14,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.lbladiciones,pos=(14,6),span=(1,3), flag= wx.ALL|wx.ALIGN_CENTER, border=5)
        self.fgs.Add(self.lblpreguntahoras,pos=(20,1),span=(2,2), flag= wx.ALL |wx.ALIGN_RIGHT | wx.ALIGN_CENTER_VERTICAL , border=5)
        self.fgs.Add(self.lblpreguntadoc,pos=(17,1),span=(2,2), flag= wx.ALL |wx.ALIGN_RIGHT | wx.ALIGN_CENTER_VERTICAL, border=5)
        self.fgs.Add(self.txtcotizacion,pos=(7,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txttipotransp,pos=(9,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txttipocont,pos=(10,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtdescargue,pos=(11,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txttiporeq,pos=(8,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtorigen,pos=(7,4),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtdestino,pos=(8,4),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtkm,pos=(9,4),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtprecio,pos=(10,4),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtrecargopeaje,pos=(11,4),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtnombreresp,pos=(7,6),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txttelresp,pos=(8,6),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtcargoresp,pos=(9,6),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtnombresiso,pos=(10,6),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txttelesiso,pos=(9,8),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtdebeinfo,pos=(7,8),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txthorasantes,pos=(8,8),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtfechaentrega,pos=(13,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtdireccion,pos=(13,4),span=(1,3), flag= wx.ALL | wx.EXPAND, border=5)
        self.fgs.Add(self.txtreferenciacont,pos=(13,8),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtnombreconduc,pos=(15,4),span=(1,2), flag= wx.ALL | wx.EXPAND, border=5)
        self.fgs.Add(self.txtcedula,pos=(15,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txttelefonoconduc,pos=(15,2),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtplaca,pos=(15,1),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.txtadiciones,pos=(15,6),span=(4,3), flag= wx.ALL| wx.EXPAND, border=5)
        self.fgs.Add(self.checkpreguntahoras_si,pos=(20,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.checkpreguntahoras_no,pos=(21,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.checkpreguntadoc_si,pos=(17,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(self.checkpreguntadoc_no,pos=(18,3),span=(1,1), flag= wx.ALL, border=5)
        self.fgs.Add(btn_finalizar,pos=(21,8),span=(1,1), flag= wx.ALL, border=5)

        self.fgs.Add(self.txtnombrecliente,pos=(11,6),span=(1,2), flag= wx.ALL| wx.EXPAND, border=5)
        self.fgs.Add(self.lblnombrecliente,pos=(11,5),span=(1,1), flag= wx.ALL, border=5)


        self.checkpreguntahoras_si.Bind(wx.EVT_CHECKBOX, self.onCheckhoras_si)
        self.checkpreguntahoras_no.Bind(wx.EVT_CHECKBOX, self.onCheckhoras_no)
        
        self.checkpreguntadoc_si.Bind(wx.EVT_CHECKBOX, self.onCheckdoc_si)
        self.checkpreguntadoc_no.Bind(wx.EVT_CHECKBOX, self.onCheckdoc_no)
        
        btn_finalizar.Bind(wx.EVT_BUTTON, self.finalizar)
        
       
        mainSizer= wx.BoxSizer(wx.VERTICAL)
        mainSizer.Add(self.fgs,0, flag=wx.ALIGN_LEFT)
        self.panel.SetSizerAndFit(mainSizer)
        
        if self.lista_valores_fila[col_debeinfo -1]==self.lista_descargue[3]:
         #   self.lbldebeinfo.Hide()
          #  self.lblhorasantes.Hide()
           # self.txtdebeinfo.Hide()
            self.txthorasantes.SetValue('N/A')
            self.lblpreguntahoras.Hide()
            self.checkpreguntahoras_si.Hide()
            self.checkpreguntahoras_no.Hide()


  
    def onCheckhoras_si(self,event):
        if self.checkpreguntahoras_no.IsChecked():
            self.checkpreguntahoras_no.SetValue(False)
            
    def onCheckhoras_no(self,event):
        if self.checkpreguntahoras_si.IsChecked():
            self.checkpreguntahoras_si.SetValue(False)
            
    def onCheckdoc_si(self,event):
        if self.checkpreguntadoc_no.IsChecked():
            self.checkpreguntadoc_no.SetValue(False)
            
    def onCheckdoc_no(self,event):
        if self.checkpreguntadoc_si.IsChecked():
            self.checkpreguntadoc_si.SetValue(False)
            
            
        
    def finalizar(self,event):
        wb_req=openpyxl.load_workbook(path_db)
        self.hist_req_sheet=wb_req['Requerimientos']
        
        
        diccionario_campos_oblig_texto={self.txtfechaentrega:'Fecha Entrega', self.txtdireccion:'Direccion Exacta', 
                                        self.txtreferenciacont:'Referencia Contenedor', self.txtreferenciacont:'Ref. Contenedor', 
                                        self.txtplaca:'Placa', self.txttelefonoconduc:'Telefono', self.txtcedula:'Cedula',
                                        self.txtnombreconduc:'Nombre Conductor'}
          
        checkbox1=[self.checkpreguntadoc_no,self.checkpreguntadoc_si]
        
        if self.validar_campos_vacios_texto(diccionario_campos_oblig_texto)==False:
            return
        if self.validar_checkbox(checkbox1,'Documentacion Completa?')==False:
            return
        
        if self.lista_valores_fila[col_debeinfo -1]==self.lista_descargue[3]:
            pass
        else:
            checkbox2=[self.checkpreguntahoras_no,self.checkpreguntahoras_si]
            if self.validar_checkbox(checkbox2,'Documentacion Enviada Horas Antes?')==False:
                return

        fechaentrega=self.txtfechaentrega.GetValue()
        direccion=self.txtdireccion.GetValue()
        referenciacont=self.txtreferenciacont.GetValue()
        nombreconduc=self.txtnombreconduc.GetValue()
        cedula=self.txtcedula.GetValue()
        telefonoconduc=self.txttelefonoconduc.GetValue()
        placa=self.txtplaca.GetValue()
        adiciones=self.txtadiciones.GetValue()
        
        if self.checkpreguntahoras_si.IsChecked():
             check_horas="Si"
        else:
            check_horas="No"
        
        if self.checkpreguntadoc_si.IsChecked():
            check_doc="Si"
        else:
            check_doc="No"
        
        
        self.hist_req_sheet.cell(row=self.nro_fila_req, column=col_fechaentrega).value=fechaentrega
        self.hist_req_sheet.cell(row=self.nro_fila_req, column=col_direccion).value=direccion
        self.hist_req_sheet.cell(row=self.nro_fila_req, column=col_referenciacont).value=referenciacont
        self.hist_req_sheet.cell(row=self.nro_fila_req, column=col_nombreconduc).value=nombreconduc
        self.hist_req_sheet.cell(row=self.nro_fila_req, column=col_cedula).value=cedula
        self.hist_req_sheet.cell(row=self.nro_fila_req, column=col_telefonoconduc).value=telefonoconduc
        self.hist_req_sheet.cell(row=self.nro_fila_req, column=col_placa).value=placa
        self.hist_req_sheet.cell(row=self.nro_fila_req, column=col_adiciones).value=adiciones
        self.hist_req_sheet.cell(row=self.nro_fila_req, column=col_preguntahoras).value=check_horas
        self.hist_req_sheet.cell(row=self.nro_fila_req, column=col_preguntadoc).value=check_doc
        
        try:
            wb_req.save(path_db)
            sgto_msgbox=wx.MessageDialog(None,'Recuerde Hacer el Seguimiento','Atencion',wx.ICON_WARNING)
            sgto_msgbox.ShowModal()
            self.Destroy()
        except:
            error_msgbox=wx.MessageDialog(None,'Error al guardar el registro en la BD. \nVerifique el el archivo de excel este cerrado y en la ruta correcta.','ERROR',wx.ICON_ERROR)
            error_msgbox.ShowModal()

    
    def validar_campos_vacios_texto(self,diccionario_campos_oblig):
        for campo in diccionario_campos_oblig:
            if len(campo.GetValue().strip()) == 0:
                error_msgbox=wx.MessageDialog(None,'Falta diligenciar el campo: ' + diccionario_campos_oblig[campo],'ERROR',wx.ICON_ERROR)
                error_msgbox.ShowModal()   
                return False
        return True
    def validar_checkbox(self,checkbox,label):
        for i in range (len(checkbox)):
            if checkbox[i].IsChecked():
                return True
        error_msgbox=wx.MessageDialog(None,'Seleccione una opcion en el campo: ' + label,'ERROR',wx.ICON_ERROR)
        error_msgbox.ShowModal()
        return False

class LogisticaPanel(wx.Panel):

    def __init__(self,parent):
        # create the panel
        wx.Panel.__init__(self, parent=parent)
        try:

            image_file = 'logo35.png'
            bmp1 = wx.Image(
                image_file, 
                wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            # image's upper left corner anchors at panel 
            # coordinates (0, 0)
            self.bitmap3 = wx.StaticBitmap(
                self, -1, bmp1, (5, 0))
            # show some image details
            #str1 = "%s  %dx%d" % (image_file, bmp1.GetWidth(),
                                  #bmp1.GetHeight()) 
            #parent.SetTitle(str1)
        except IOError:
            print ("Image file %s not found")
            raise SystemExit        


class ww_remision11(wx.Frame):
    
    def __init__(self,parent):
        
        self.wb_req=openpyxl.load_workbook(path_db)
        self.hist_req_sheet=self.wb_req['Requerimientos']
        
        wx.Frame.__init__(self, None, wx.ID_ANY, "Centro Logistico", size=(270, 250),style=wx.DEFAULT_FRAME_STYLE & ~(wx.RESIZE_BORDER | wx.MAXIMIZE_BOX))  
        self.SetBackgroundColour(secondary_color)
        self.Center()
        try:
            
            #image_file = 'CINCO CONSULTORES.jpg'
            #bmp1 = wx.Image(
                #image_file, 
                #wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            #self.panel = wx.StaticBitmap(
                #self, -1, bmp1, (0, 0)
            self.panel=wx.Panel(self)
            panel_font= wx.Font(10, wx.DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Normal")
            self.panel.SetBackgroundColour(secondary_color)
            self.panel.SetFont(panel_font)

        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
        
        ico = wx.Icon('Cont.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        self.fgs= wx.GridBagSizer(0,0)
        
        title_font= wx.Font(11, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL,underline=False,faceName="Folks-Normal")
       
        self.lbltitle =wx.StaticText(self.panel, label='Ingrese Numero de Requerimiento:')
        self.lbltitle.SetFont(title_font)
        self.lbltitle.SetBackgroundColour(secondary_color)
        self.lbltitle.SetForegroundColour(principal_color)
        self.fgs.Add(self.lbltitle,pos=(2,1),span=(1,3), flag=wx.ALL | wx.ALIGN_CENTER, border=5)

        self.txtreq = wx.TextCtrl(self.panel)
        self.fgs.Add(self.txtreq , pos=(4,1),span=(1,3), flag= wx.ALL| wx.ALIGN_CENTER, border=5)
        
        btn_aceptar = wx.Button(self.panel, id=wx.ID_OK, label="Aceptar",size=(-1,-1))
        self.fgs.Add(btn_aceptar, pos=(6,1),span=(1,3), flag= wx.ALL | wx.ALIGN_CENTER, border=0)
        btn_aceptar.Bind(wx.EVT_BUTTON, self.open_remision22)

        mainSizer= wx.BoxSizer(wx.VERTICAL)
        mainSizer.Add(self.fgs,0, flag=wx.ALIGN_LEFT)
        self.panel.SetSizerAndFit(mainSizer)
        
    #-------------Button Functions-----------------#            
    def open_remision22(self, event):
        
        self.lista_nro_req=[]
        
        for cell in self.hist_req_sheet['A']:
            if cell.value !=None:
                self.lista_nro_req.append(cell.value)
        global req_selec
        
        try:
            req_selec=int(self.txtreq.GetValue())
        except:
            error_msgbox=wx.MessageDialog(None,'Numero de Requerimiento No Encontrado','ERROR',wx.ICON_ERROR)
            error_msgbox.ShowModal()
            return
        
        if req_selec in self.lista_nro_req:
            index_req=self.lista_nro_req.index(req_selec)
            self.crear_remision(index_req)
            self.Destroy()
        else:
            error_msgbox=wx.MessageDialog(None,'Numero de Requerimiento No Encontrado','ERROR',wx.ICON_ERROR)
            error_msgbox.ShowModal()
            
    def crear_remision(self,index_req):
        
        wb_remision=openpyxl.load_workbook(path_remision)
        sheet_remision=wb_remision['Remision']
        
        
        rows=[]
        for row in self.hist_req_sheet.iter_rows(min_row=(index_req+1), max_row=(index_req+1)):
            lbls=[]
            for cell in row:
                lbls.append(cell.value)
            rows.append(lbls)
        
        
        nro_req=str(rows[0][col_requerimiento_auto-1])
        tipo=rows[0][col_tipo_req-1]
        fecha=datetime.today().strftime('%d-%m-%Y')
        remite='CONTENEDORES DE ANTIOQUIA S.A.S'
        cliente=rows[0][col_nombrecliente-1]
        destino=rows[0][col_destino-1]
        origen=rows[0][col_origen-1]
        direccion=rows[0][col_direccion-1]
        responsable=rows[0][col_nombreresponsable-1]
        celular=rows[0][col_telefono_resp-1]
        placa=rows[0][col_placa-1]
        conductor=rows[0][col_nombreconduc-1]
        cedula=rows[0][col_cedula-1]
        telefono=rows[0][col_telefonoconduc-1]
        adiciones=rows[0][col_adiciones-1]
        nro_interno=rows[0][col_referenciacont-1]
        
        sheet_remision['K9']=nro_req
        sheet_remision['AC9']=tipo
        sheet_remision['H12']=fecha
        sheet_remision['H13']=remite
        sheet_remision['H14']=cliente
        sheet_remision['H15']=destino
        sheet_remision['H16']=origen
        sheet_remision['H17']=direccion
        sheet_remision['H18']=responsable
        sheet_remision['H19']=celular
        sheet_remision['H21']=placa
        sheet_remision['H22']=conductor
        sheet_remision['H23']=cedula
        sheet_remision['H24']=telefono
        sheet_remision['H27']=adiciones
        sheet_remision['H35']=nro_interno
        
        if len(nro_req)==3:
            str_nro_req=nro_req
        elif len(nro_req)==2:
            str_nro_req='0'+str(nro_req)
        elif len(nro_req)==1:
            str_nro_req='00'+str(nro_req)
         
        año=datetime.today().strftime('%Y')
        fecha_remision=año[2:4]
        
        sheet_remision.cell(row=9, column=11)._style=deepcopy(sheet_remision['B9']._style)
        new_path=path_remision_nuevas + 'Remision No ' + str_nro_req + '-' + str(fecha_remision) +' '+ cliente.upper() + ' .xlsx'
        wb_remision.save(new_path)   

class ww_configuracion(wx.Frame):   
    
    def __init__(self,parent):
   

        
        wx.Frame.__init__(self, None, wx.ID_ANY, "Contenedores de Antioquia - Centro Logistico", size=(250, 250))  
        
        try:
            
            #image_file = 'CINCO CONSULTORES.jpg'
            #bmp1 = wx.Image(
                #image_file, 
                #wx.BITMAP_TYPE_ANY).ConvertToBitmap()
            
            #self.panel = wx.StaticBitmap(
                #self, -1, bmp1, (0, 0)
            self.panel=wx.Panel(self)
            self.panel.SetBackgroundColour(secondary_color)

        except IOError:
            print ("Image file %s not found"  )
            raise SystemExit
        
        ico = wx.Icon('Cont.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ico)
        self.fgs= wx.GridBagSizer(0,0)
        
        title_font= wx.Font(10, wx.FONTFAMILY_DECORATIVE, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
       
        self.lbltitle =wx.StaticText(self.panel, label='Ingrese Contraseña')
        self.lbltitle.SetFont(title_font)
        self.lbltitle.SetBackgroundColour(secondary_color)
        self.lbltitle.SetForegroundColour(principal_color)
        self.fgs.Add(self.lbltitle,pos=(2,1),span=(1,3), flag=wx.ALL | wx.ALIGN_CENTER, border=5)

        self.txtpass = wx.TextCtrl(self.panel, style= wx.TE_PASSWORD)
        self.fgs.Add(self.txtpass , pos=(3,1),span=(1,3), flag= wx.ALL | wx.EXPAND, border=5)
        
        btn_aceptar = wx.Button(self.panel, id=wx.ID_ANY, label="Aceptar",size=(-1,-1))
        self.fgs.Add(btn_aceptar, pos=(6,1),span=(1,3), flag= wx.ALL | wx.ALIGN_CENTER, border=0)
        btn_aceptar.Bind(wx.EVT_BUTTON, self.onBtn_aceptar)

        mainSizer= wx.BoxSizer(wx.VERTICAL)
        mainSizer.Add(self.fgs,0, flag=wx.ALIGN_CENTER)
        self.panel.SetSizerAndFit(mainSizer)
        
    #-------------Button Functions-----------------#            
    def onBtn_aceptar(self, event):
        self.config_sheet=wb_listas['Config']
        
        self.Destroy()
        #ww_nuevo_requerimiento12(parent=self.panel).Show()

    #-------------Button Functions-----------------# 
    

       

class MyApp(wx.App):
    def OnInit(self):
        self.frame= MyFrame()
        self.frame.Show()
        return True       

# Run the program     
app=MyApp()
app.MainLoop()
del app
            
            